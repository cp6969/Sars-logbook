"""Shared SARS export data-prep, used by BOTH the HTML print report and
the Excel export so the two outputs can never arithmetically drift apart
-- only their presentation differs.

Odometer reconciliation, the real design nuance here: SARS wants literal
odometer readings, but the phone can only reliably give GPS-derived
distance automatically. The ledger below starts at the vehicle's own
tax-year-opening odometer and accumulates each trip's distance_km in
chronological order -- UNLESS a trip has a REAL manual odometer_open
and/or odometer_close entered, in which case that real value takes
precedence and re-anchors the running total from that point forward. A
user who occasionally types their real odometer reading gets a
self-correcting, more defensible logbook over time; a user who never
types one still gets a fully GPS-derived logbook, which is still a widely
accepted proxy.
"""
from datetime import date

from app.trips import list_trips


def tax_year_bounds(reference_date: date, start_month: int, start_day: int):
    """Given any date, returns (start, end) of the SA tax year that date
    falls in. `end` is exclusive (the start of the NEXT tax year)."""
    year = reference_date.year
    tax_start = date(year, start_month, start_day)
    if reference_date < tax_start:
        tax_start = date(year - 1, start_month, start_day)
    tax_end = date(tax_start.year + 1, start_month, start_day)
    return tax_start, tax_end


def compute_odometer_ledger(opening_odometer, trips: list[dict]) -> list[dict]:
    """`trips` must be ordered by started_at ascending. Returns a new list
    of dicts, each trip augmented with 'computed_odometer_open' and
    'computed_odometer_close' -- see the module docstring for the
    anchoring rule."""
    ledger = []
    running = float(opening_odometer)
    for trip in trips:
        distance = float(trip["distance_km"] or 0)
        odo_open = trip.get("odometer_open")
        odo_close = trip.get("odometer_close")

        open_val = float(odo_open) if odo_open is not None else running
        close_val = float(odo_close) if odo_close is not None else open_val + distance

        ledger.append(
            {**trip, "computed_odometer_open": round(open_val, 1), "computed_odometer_close": round(close_val, 1)}
        )
        running = close_val
    return ledger


def current_odometer(vehicle: dict, trips: list[dict]) -> float:
    """The vehicle's current computed odometer reading -- the closing
    value of the most recent trip in the ledger, or the tax-year opening
    reading if there are no trips yet."""
    ledger = compute_odometer_ledger(vehicle["tax_year_opening_odometer"], trips)
    if not ledger:
        return round(float(vehicle["tax_year_opening_odometer"]), 1)
    return ledger[-1]["computed_odometer_close"]


def build_report_data(vehicle: dict, start_date: date, end_date: date) -> dict:
    """The ledger is always recomputed from the vehicle's own
    tax-year-opening odometer forward through EVERY trip up to end_date
    (not just ones inside start_date) -- otherwise a report run mid-year
    for e.g. "just this month" would show a wrong opening odometer for
    that month, since the running total has to carry forward from the
    true start of the tax year, not from an arbitrary report-window
    start."""
    all_trips_to_date = list_trips(vehicle["id"], start_date=None, end_date=end_date)
    ledger = compute_odometer_ledger(vehicle["tax_year_opening_odometer"], all_trips_to_date)

    rows = [row for row in ledger if row["ended_at"] is not None and row["started_at"].date() >= start_date]

    business_km = sum(float(r["distance_km"] or 0) for r in rows if r["category"] == "business")
    private_km = sum(float(r["distance_km"] or 0) for r in rows if r["category"] == "private")
    unclassified_km = sum(float(r["distance_km"] or 0) for r in rows if r["category"] is None)

    if rows:
        opening_odometer = rows[0]["computed_odometer_open"]
        closing_odometer = rows[-1]["computed_odometer_close"]
    elif ledger:
        opening_odometer = closing_odometer = ledger[-1]["computed_odometer_close"]
    else:
        opening_odometer = closing_odometer = float(vehicle["tax_year_opening_odometer"])

    return {
        "vehicle": vehicle,
        "start_date": start_date,
        "end_date": end_date,
        "rows": rows,
        "opening_odometer": round(opening_odometer, 1),
        "closing_odometer": round(closing_odometer, 1),
        "total_km": round(business_km + private_km + unclassified_km, 1),
        "business_km": round(business_km, 1),
        "private_km": round(private_km, 1),
        "unclassified_km": round(unclassified_km, 1),
    }


def build_excel_report(data: dict) -> bytes:
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "SARS Logbook"
    bold = Font(bold=True)

    ws["A1"] = "SARS Business Travel Logbook"
    ws["A1"].font = Font(bold=True, size=14)

    vehicle = data["vehicle"]
    summary_rows = [
        ("Vehicle registration", vehicle["registration"]),
        ("Make / model", f"{vehicle.get('make') or ''} {vehicle.get('model') or ''}".strip() or "-"),
        ("Period", f"{data['start_date'].strftime('%d/%m/%Y')} - {data['end_date'].strftime('%d/%m/%Y')}"),
        ("Opening odometer", data["opening_odometer"]),
        ("Closing odometer", data["closing_odometer"]),
        ("Total km travelled", data["total_km"]),
        ("Business km", data["business_km"]),
        ("Private km", data["private_km"]),
    ]
    for i, (label, value) in enumerate(summary_rows, start=3):
        ws.cell(row=i, column=1, value=label).font = bold
        ws.cell(row=i, column=2, value=value)

    header_row = 3 + len(summary_rows) + 2
    headers = ["Date", "Opening km", "Closing km", "Total km", "Category", "Details of trip"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=header).font = bold

    for i, trip in enumerate(data["rows"], start=header_row + 1):
        ws.cell(row=i, column=1, value=trip["started_at"].strftime("%d/%m/%Y"))
        ws.cell(row=i, column=2, value=trip["computed_odometer_open"])
        ws.cell(row=i, column=3, value=trip["computed_odometer_close"])
        ws.cell(row=i, column=4, value=float(trip["distance_km"] or 0))
        ws.cell(row=i, column=5, value=(trip["category"] or "").title())
        ws.cell(row=i, column=6, value=trip["purpose"] or "")

    for col_letter, width in zip("ABCDEF", [14, 12, 12, 10, 12, 45]):
        ws.column_dimensions[col_letter].width = width

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
