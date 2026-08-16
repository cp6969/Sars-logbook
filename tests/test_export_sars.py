from datetime import date, datetime, timezone
from io import BytesIO

from openpyxl import load_workbook

from app import trips as trips_module
from app.export_sars import build_excel_report, build_report_data


def test_html_and_excel_reconcile(test_vehicle):
    start1 = trips_module.start_trip(test_vehicle["id"], "u1", datetime(2026, 3, 10, tzinfo=timezone.utc), -29.0, 30.0)
    trips_module.end_trip(start1["id"], datetime(2026, 3, 10, 1, tzinfo=timezone.utc), -29.1, 30.0)
    trips_module.classify_trip(start1["id"], "business", "Client visit")

    start2 = trips_module.start_trip(test_vehicle["id"], "u2", datetime(2026, 3, 11, tzinfo=timezone.utc), -29.1, 30.0)
    trips_module.end_trip(start2["id"], datetime(2026, 3, 11, 1, tzinfo=timezone.utc), -29.0, 30.0)
    trips_module.classify_trip(start2["id"], "private", "")

    data = build_report_data(test_vehicle, date(2026, 3, 1), date(2027, 3, 1))
    assert data["business_km"] > 0
    assert data["private_km"] > 0
    assert round(data["business_km"] + data["private_km"], 1) == data["total_km"]
    assert round(data["closing_odometer"] - data["opening_odometer"], 1) == data["total_km"]

    excel_bytes = build_excel_report(data)
    wb = load_workbook(BytesIO(excel_bytes))
    ws = wb.active
    # header row is 3 (summary start) + 8 summary rows + 2 blank = row 13
    assert ws["B8"].value == data["total_km"]
    assert ws["B9"].value == data["business_km"]
    assert ws["B10"].value == data["private_km"]


def test_report_excludes_trips_outside_the_window(test_vehicle):
    trip = trips_module.start_trip(test_vehicle["id"], "u3", datetime(2026, 3, 10, tzinfo=timezone.utc), -29.0, 30.0)
    trips_module.end_trip(trip["id"], datetime(2026, 3, 10, 1, tzinfo=timezone.utc), -29.1, 30.0)
    trips_module.classify_trip(trip["id"], "business", "In window")

    trip2 = trips_module.start_trip(test_vehicle["id"], "u4", datetime(2026, 6, 1, tzinfo=timezone.utc), -29.0, 30.0)
    trips_module.end_trip(trip2["id"], datetime(2026, 6, 1, 1, tzinfo=timezone.utc), -29.1, 30.0)
    trips_module.classify_trip(trip2["id"], "business", "Out of window")

    data = build_report_data(test_vehicle, date(2026, 3, 1), date(2026, 4, 1))
    purposes = [r["purpose"] for r in data["rows"]]
    assert "In window" in purposes
    assert "Out of window" not in purposes
